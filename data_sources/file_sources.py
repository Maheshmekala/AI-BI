"""File-based data sources: CSV, Excel, PDF — all DuckDB-backed."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

from data_sources.base import DataSource, Dataset
from sql_engine import SqlEngine, random_table_name


class CSVSource(DataSource):
    source_type = "csv"
    display_name = "CSV File"

    def __init__(self, config: dict[str, Any] | None = None) -> None:
        super().__init__(config)
        self.file_path: Optional[Path] = None

    def connect(self) -> bool:
        path = self.config.get("file_path") or self.config.get("path")
        if not path:
            raise ValueError("CSVSource requires 'file_path' in config")
        self.file_path = Path(path)
        return self.file_path.exists()

    def load(self) -> list[Dataset]:
        sep = self.config.get("sep", ",")
        header = self.config.get("header", True)
        table_name = random_table_name("csv")

        # DuckDB reads CSV directly — zero-copy, no pandas
        path_str = str(self.file_path).replace("\\", "/")
        self.engine.ingest_csv(path_str, table_name, delim=sep, header=header)

        name = self.file_path.stem.replace(" ", "_").replace("-", "_")
        return [
            Dataset(
                name=name,
                table_name=table_name,
                sql_engine=self.engine,
                source_type="csv",
            )
        ]


class ExcelSource(DataSource):
    source_type = "excel"
    display_name = "Excel File"

    def __init__(self, config: dict[str, Any] | None = None) -> None:
        super().__init__(config)
        self.file_path: Optional[Path] = None

    def connect(self) -> bool:
        path = self.config.get("file_path") or self.config.get("path")
        if not path:
            raise ValueError("ExcelSource requires 'file_path' in config")
        self.file_path = Path(path)
        return self.file_path.exists()

    def load(self) -> list[Dataset]:
        import openpyxl

        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        datasets = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)

            # First row = headers
            try:
                headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(rows_iter))]
            except StopIteration:
                continue

            # Clean up headers
            headers = [h.replace(" ", "_").replace("-", "_") for h in headers]

            # Read data rows
            data_rows = []
            for row in rows_iter:
                data_rows.append(list(row))
                if len(data_rows) >= 50000:  # Safety limit per sheet
                    break

            if not data_rows:
                continue

            # Create DuckDB table
            table_name = random_table_name("xls")
            col_defs = ", ".join(f'"{h}" VARCHAR' for h in headers)
            self.engine.execute(f"CREATE TABLE {table_name} ({col_defs})")

            # Insert rows in batches
            batch_size = 1000
            for i in range(0, len(data_rows), batch_size):
                batch = data_rows[i:i + batch_size]
                placeholders = ", ".join(["?" for _ in headers])
                values_sql = ", ".join([f"({placeholders})" for _ in batch])
                flat_params = []
                for row in batch:
                    flat_params.extend(str(v) if v is not None else None for v in row)
                try:
                    self.engine.execute(
                        f"INSERT INTO {table_name} VALUES {values_sql}",
                        params=flat_params,
                    )
                except Exception:
                    # Try inserting one by one on failure
                    for row in batch:
                        try:
                            self.engine.execute(
                                f"INSERT INTO {table_name} VALUES ({placeholders})",
                                params=[str(v) if v is not None else None for v in row],
                            )
                        except Exception:
                            pass

            ds_name = f"{self.file_path.stem}_{sheet_name}".replace(" ", "_").replace("-", "_")
            datasets.append(
                Dataset(
                    name=ds_name,
                    table_name=table_name,
                    sql_engine=self.engine,
                    source_type="excel",
                )
            )

        wb.close()
        return datasets


class PDFSource(DataSource):
    source_type = "pdf"
    display_name = "PDF File"

    def __init__(self, config: dict[str, Any] | None = None) -> None:
        super().__init__(config)
        self.file_path: Optional[Path] = None

    def connect(self) -> bool:
        path = self.config.get("file_path") or self.config.get("path")
        if not path:
            raise ValueError("PDFSource requires 'file_path' in config")
        self.file_path = Path(path)
        return self.file_path.exists()

    def _extract_tables(self) -> list[tuple[str, list[dict[str, Any]]]]:
        """Extract tabular data from PDF pages as list-of-dicts."""
        import pdfplumber

        tables = []
        with pdfplumber.open(self.file_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                extracted = page.extract_table()
                if extracted and len(extracted) > 1:
                    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(extracted[0])]
                    rows = []
                    for row in extracted[1:]:
                        cleaned = [str(c).strip() if c else "" for c in row]
                        if any(c for c in cleaned):
                            rows.append(dict(zip(headers, cleaned)))
                    if rows:
                        tables.append((f"table_page_{page_num + 1}", rows, headers))
        return tables  # type: ignore

    def _extract_text(self) -> list[dict[str, Any]] | None:
        """Fallback: extract raw text into rows."""
        import pdfplumber

        lines = []
        with pdfplumber.open(self.file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    lines.extend(text.split("\n"))
        if not lines:
            return None
        return [{"content": l.strip()} for l in lines if l.strip()]

    def load(self) -> list[Dataset]:
        datasets = []

        # First try tables
        tables = self._extract_tables()  # type: ignore
        for name, rows, headers in tables:
            if not rows:
                continue
            table_name = random_table_name("pdf")
            col_defs = ", ".join(f'"{h}" VARCHAR' for h in headers)
            self.engine.execute(f"CREATE TABLE {table_name} ({col_defs})")

            # Batch insert
            batch_size = 500
            placeholders = ", ".join(["?" for _ in headers])
            for i in range(0, len(rows), batch_size):
                batch = rows[i:i + batch_size]
                values_sql = ", ".join([f"({placeholders})" for _ in batch])
                flat_params = []
                for row in batch:
                    flat_params.extend(str(row.get(h, "")) for h in headers)
                try:
                    self.engine.execute(
                        f"INSERT INTO {table_name} VALUES {values_sql}",
                        params=flat_params,
                    )
                except Exception:
                    for row in batch:
                        try:
                            self.engine.execute(
                                f"INSERT INTO {table_name} VALUES ({placeholders})",
                                params=[str(row.get(h, "")) for h in headers],
                            )
                        except Exception:
                            pass

            ds_name = f"{self.file_path.stem}_{name}".replace(" ", "_").replace("-", "_")
            datasets.append(
                Dataset(
                    name=ds_name,
                    table_name=table_name,
                    sql_engine=self.engine,
                    source_type="pdf",
                    description=f"Extracted table from {name}",
                )
            )

        # Fallback text extraction
        if not datasets:
            text_rows = self._extract_text()
            if text_rows:
                table_name = random_table_name("pdf")
                self.engine.execute(f"CREATE TABLE {table_name} (\"content\" VARCHAR)")
                self.engine.ingest_rows(
                    table_name, ["content"],
                    [[r["content"]] for r in text_rows],
                )
                name = self.file_path.stem.replace(" ", "_").replace("-", "_")
                datasets.append(
                    Dataset(
                        name=name,
                        table_name=table_name,
                        sql_engine=self.engine,
                        source_type="pdf",
                        description="Extracted text content",
                    )
                )

        return datasets
